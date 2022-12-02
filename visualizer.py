from typing import Dict, List, Tuple
from agent import Agent
import numpy as np
import math
from matplotlib import pyplot as plt
from matplotlib.animation import FuncAnimation
import matplotlib.patches as patches
import openpyxl


class Visualizer:

    def __init__(self,
                 excel_map_file: str,
                 solution: Dict[Agent, np.ndarray],
                 step_div=10):
        self.load_excel_map(excel_map_file)
        self.solution = solution
        self.traject: Dict[Agent, np.ndarray] = dict()
        self.step_div = step_div
        self.path_steps = 0
        self.traj_steps = 0
        self._make_traject()
        self._set_steps()
    
    def load_excel_map(self, excel_map_file):
        wb = openpyxl.load_workbook(excel_map_file)
        map_sh = wb['map']
        self.grid_size_x = map_sh.max_row
        self.grid_size_y = map_sh.max_column
        self.static_obstacles = []
        for row in range(1, map_sh.max_row + 1):
            for col in range(1, map_sh.max_column + 1):
                if map_sh.cell(row, col).value == '@':
                    x = row - 1
                    y = col - 1
                    self.static_obstacles.append([x, y])

    def _set_steps(self):
        for path in self.solution.values():
            self.path_steps = max(self.path_steps, len(path))
        self.traj_steps = int(self.path_steps * self.step_div + 1)

    def _make_traject(self):
        for agent, path in self.solution.items():
            traj_x = np.array([path[0][0]])
            traj_y = np.array([path[0][1]])
            for timestep in range(len(path) - 1):
                now_pos = path[timestep]
                next_pos = path[timestep + 1]
                for i in range(self.step_div):
                    i_pos = now_pos + (next_pos - now_pos) / self.step_div * (i + 1)
                    traj_x = np.append(traj_x, i_pos[0])
                    traj_y = np.append(traj_y, i_pos[1])
            self.traject[agent] = np.stack([traj_x, traj_y], 1)

    def plot(self, save=False):
        fig = plt.figure()
        ax = fig.add_subplot(111, aspect=1)
        ax.invert_yaxis()

        def plot_one_step(f):
            # clear pre step artists
            ax.cla()
            ax.set_xlim(-1, self.grid_size_y)
            ax.set_ylim(-1, self.grid_size_x)
            ax.invert_yaxis()
            # set grid
            for x in range(self.grid_size_y + 1):
                ax.plot([x - 0.5, x - 0.5], [0 - 0.5, self.grid_size_x - 0.5], color="black")
            for y in range(self.grid_size_x + 1):
                ax.plot([0 - 0.5, self.grid_size_y - 0.5], [y - 0.5, y - 0.5], color="black")
            # plot static obstacle
            for obstacle in self.static_obstacles:
                obsy = obstacle[0] - 0.5
                obsx = obstacle[1] - 0.5
                r = patches.Rectangle(xy=(obsx, obsy), width=1, height=1, color="black")
                ax.add_patch(r)
            # plot agents
            for agent, pos in self.traject.items():
                # plot agents            
                if f < len(pos):
                    posy = pos[f][0]
                    posx = pos[f][1]
                    disty = pos[-1][0]
                    distx = pos[-1][1]
                    c = patches.Circle(xy=(posx, posy), radius=0.3, color="blue")
                    cg = patches.Circle(xy=(distx, disty), radius=0.3, color="green", alpha=0.4)
                    ax.text(posx, posy, str(agent.id), va="center", ha="center", fontsize=8, color="white")
                    ax.text(posx, posy, str(agent.id), va="center", ha="center", fontsize=8, color="white")
                    ax.add_patch(c)
                    ax.add_patch(cg)
                else:
                    disty = pos[-1][0]
                    distx = pos[-1][1]
                    c = patches.Circle(xy=(distx, disty), radius=0.3, color="red")
                    ax.text(distx, disty, str(agent.id), va="center", ha="center", fontsize=8, color="white")
                    ax.add_patch(c)
        interval_time = 1000/self.step_div
        anim = FuncAnimation(fig, plot_one_step, frames=self.traj_steps, interval=interval_time)
        if save:
            anim.save('./anim/anim.mp4', writer="ffmpeg")
        plt.show()


if __name__ == '__main__':
    agent_1 = Agent(1)
    agent_2 = Agent(2)
    excel_map_file = "./map/map.xlsx"
    path_1 = np.array([[1, 2], [2, 2], [3, 2], [3, 3]])
    path_2 = np.array([[2, 2], [2, 3], [3, 3], [4, 3], [5, 3]])
    solution = {agent_1: path_1, agent_2: path_2}

    visualizer = Visualizer(excel_map_file, solution)
    print(visualizer.traject)
    visualizer.plot()

