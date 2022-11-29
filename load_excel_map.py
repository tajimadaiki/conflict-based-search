import openpyxl
from typing import List

def load_excel_map(map_file):
        wb = openpyxl.load_workbook(map_file)
        map_sh = wb['map']
        map: List[List[str]] = []
        for row in range(1, map_sh.max_row + 1):
            map_row = []
            for col in range(1, map_sh.max_column + 1):
                map_row.append(map_sh.cell(row, col).value)
            map.append(map_row)
            print(map)


if __name__ == "__main__":
    map_file = "./map/map.xlsx"
    load_excel_map(map_file)
