from typing import List, Tuple, Dict
import numpy as np
import openpyxl


class NeighbourTable:

    def __init__(self,
                 excel_map_file: str):
        # load map file
        self.load_excel_map(excel_map_file)
        # create neighbours table
        self.table: Dict[Tuple[int,int]: str] = dict()
        for x in range(self.grid_size_x):
            for y in range(self.grid_size_y):
                directions = []
                neighbours = []
                value = self.map[x][y]
                if len(value) == 1:
                    directions = [(0, 0), (1, 0), (-1, 0), (0, 1), (0, -1)]
                else:
                    directions.append((0, 0))
                    for c in value[1:]:
                        if c == 'r': directions.append((0, -1))
                        if c == 'u': directions.append((-1, 0))
                        if c == 'l': directions.append((0, 1))
                        if c == 'd': directions.append((1, 0))
                
                for dx, dy in directions:
                    nx = x+ dx
                    ny = y + dy
                    if 0 <= nx < self.grid_size_x and 0 <= ny < self.grid_size_y:
                        if not (self.map[nx][ny] == '@'):
                            neighbours.append(np.array([nx, ny]))
                self.table[(x, y)] = np.array(neighbours)             

    def neighbours(self, pos: np.ndarray) -> np.ndarray:
        return self.table[tuple(pos)]
    
    def is_obstacle(self, pos: np.ndarray) -> bool:
        x = pos[0]
        y = pos[1]
        return self.map[x][y] == '@'

    def load_excel_map(self, map_file):
        wb = openpyxl.load_workbook(map_file)
        map_sh = wb['map']
        self.map: List[List[str]] = []
        self.grid_size_x = map_sh.max_row
        self.grid_size_y = map_sh.max_column
        for row in range(1, map_sh.max_row + 1):
            map_row = []
            for col in range(1, map_sh.max_column + 1):
                map_row.append(map_sh.cell(row, col).value)
            self.map.append(map_row)

if __name__ == "__main__":
    map_file = "./map/map.xlsx"
    neighbour = NeighbourTable(map_file)
    max_x = neighbour.grid_size_x - 1
    max_y = neighbour.grid_size_y - 1
    print(max_x, max_y)
    print(neighbour.map[12][4])
    pos1 = np.array([24, 35])
    pos2 = np.array([4, 3])
    print(neighbour.neighbours(pos1))
    print(neighbour.is_obstacle(pos2))
