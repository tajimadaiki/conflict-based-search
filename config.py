import openpyxl
from typing import List
from agent import Agent

class Config:
    def __init__(self):
        self.agents: List[Agent] = []
        self.agents_num = int()
        self.agents_id: List[str] = []
        self.init_pos = dict() # agent_id: (x, y)
        self.map: List[List[str]] = []
        self.static_obstacles = []
        self.endpoints = dict()
        self.chargers = dict()
        self.grid_size_x = int()
        self.grid_size_y = int()
        
    def load_from_xlsx(self, config_file: str):
        self._wb = openpyxl.load_workbook(config_file)
        self._load_agent()
        self._load_map() 
    
    # load work sheet 'agents'
    def _load_agent(self):
        agents_ws = self._wb['agents']
        map_ws = self._wb['map']
        keys = dict()
        self.agents_num = agents_ws.max_row - 1
        for row in range(1, agents_ws.max_row + 1):
            agent_id = ''
            for col in range(1, agents_ws.max_column + 1):
                if row == 1:
                    key = str(agents_ws.cell(row, col).value)
                    keys[col] = key
                else:
                    value = str(agents_ws.cell(row, col).value)
                    if keys[col] == 'id': 
                        agent_id = value
                        self.agents_id.append(agent_id)
                        agent = Agent(agent_id)
                        self.agents.append(agent)
                    if keys[col] == 'init_pos':
                        x = map_ws[value].row - 1
                        y = map_ws[value].column - 1
                        self.init_pos[agent_id] = (x, y)

    # load work sheet 'map' and 'name'
    def _load_map(self):
        map_ws = self._wb['map']
        name_ws = self._wb['name']
        self.grid_size_x = map_ws.max_row
        self.grid_size_y = map_ws.max_column
        for row in range(1, map_ws.max_row + 1):
            map_row = []
            for col in range(1, map_ws.max_column + 1):
                x = row - 1
                y = col - 1
                # record map
                map_row.append(map_ws.cell(row, col).value)
                name = name_ws.cell(row, col).value
                # record static obstacles
                if map_ws.cell(row, col).value == '@':
                    self.static_obstacles.append([x, y])
                # record endpoints
                if map_ws.cell(row, col).value[0] == 'e':
                    self.endpoints[name] = [x, y]
                # record chargers
                if map_ws.cell(row, col).value[0] == 'c':
                    self.chargers[name] = [x, y]
            self.map.append(map_row)


if __name__ == "__main__":
    config = Config()
    config_file = "./config/config.xlsx"
    config.load_from_xlsx(config_file)
    print(config.agents)
    print(config.agents_num)
    print(config.agents_id)
    print(config.endpoints)
    print(config.chargers)
    print(config.init_pos['1'])

